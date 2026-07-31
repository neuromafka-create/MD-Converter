"""XLSX → Markdown structural converter."""

from __future__ import annotations

from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.config import MAX_GFM_COLUMNS
from app.converters.base import BaseConverter
from app.core.optimize import optimize_markdown
from app.models import ConvertOptions, ConvertResult, FileJob
from app.utils.markdown import gfm_table, make_frontmatter, normalize_cell_text
from app.utils.paths import output_path_for


class XlsxConverter(BaseConverter):
    def convert(
        self,
        job: FileJob,
        output_dir: Path,
        options: ConvertOptions,
    ) -> ConvertResult:
        source = job.source
        bytes_in = source.stat().st_size

        try:
            # data_only=True uses cached calculated values (no Excel engine needed).
            wb = load_workbook(str(source), read_only=True, data_only=True)
        except Exception as exc:  # noqa: BLE001
            return ConvertResult(
                source=source,
                ok=False,
                error=f"XLSX open error: {exc}",
                bytes_in=bytes_in,
            )

        try:
            sheet_payloads: list[tuple[str, str]] = []
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                body = self._sheet_to_markdown(ws, options)
                if not body.strip():
                    continue
                sheet_payloads.append((sheet_name, body))
        finally:
            wb.close()

        if not sheet_payloads:
            # Empty workbook still produces a minimal note.
            sheet_payloads = [("", "_Empty workbook_\n")]

        outputs: list[Path] = []
        bytes_out = 0

        if options.excel_one_file_per_sheet and len(sheet_payloads) >= 1 and sheet_payloads[0][0]:
            for sheet_name, body in sheet_payloads:
                md = self._build_md(
                    title=f"{source.stem} — {sheet_name}",
                    source=source,
                    body=f"## {sheet_name}\n\n{body}",
                    options=options,
                    extra={"sheet": sheet_name},
                )
                dest = output_path_for(
                    job.source,
                    job.relative,
                    output_dir,
                    options,
                    suffix=sheet_name,
                )
                dest.parent.mkdir(parents=True, exist_ok=True)
                dest.write_text(md, encoding="utf-8")
                outputs.append(dest)
                bytes_out += dest.stat().st_size
        else:
            sections: list[str] = []
            for sheet_name, body in sheet_payloads:
                if sheet_name:
                    sections.append(f"## {sheet_name}\n\n{body}")
                else:
                    sections.append(body)
            combined = "\n\n".join(sections)
            md = self._build_md(
                title=source.stem,
                source=source,
                body=combined,
                options=options,
            )
            dest = output_path_for(job.source, job.relative, output_dir, options)
            dest.parent.mkdir(parents=True, exist_ok=True)
            dest.write_text(md, encoding="utf-8")
            outputs.append(dest)
            bytes_out = dest.stat().st_size

        return ConvertResult(
            source=source,
            outputs=outputs,
            ok=True,
            bytes_in=bytes_in,
            bytes_out=bytes_out,
        )

    def _build_md(
        self,
        *,
        title: str,
        source: Path,
        body: str,
        options: ConvertOptions,
        extra: dict[str, Any] | None = None,
    ) -> str:
        parts: list[str] = []
        if options.yaml_frontmatter:
            parts.append(
                make_frontmatter(
                    title=title,
                    source=source,
                    doc_type="xlsx",
                    extra=extra,
                )
            )
        parts.append(f"# {title}\n")
        parts.append(body)
        return optimize_markdown(
            "\n".join(parts),
            collapse_blank_lines=options.collapse_blank_lines,
        )

    def _sheet_to_markdown(self, ws: Worksheet, options: ConvertOptions) -> str:
        rows = self._read_rows(ws, skip_empty=options.skip_empty_excel_rows)
        if not rows:
            return "_Empty sheet_\n"

        # Trim trailing empty columns across the whole sheet.
        rows = _trim_trailing_columns(rows)
        if not rows:
            return "_Empty sheet_\n"

        width = max(len(r) for r in rows)

        if width > MAX_GFM_COLUMNS:
            return self._wide_as_blocks(rows)

        # Use first row as header if it looks non-empty / label-like.
        header = True
        if not any(c.strip() for c in rows[0]):
            header = False
        return gfm_table(rows, header=header) + "\n"

    def _read_rows(self, ws: Worksheet, *, skip_empty: bool) -> list[list[str]]:
        rows: list[list[str]] = []
        for row in ws.iter_rows(values_only=True):
            cells = [normalize_cell_text(_coerce_value(v)) for v in row]
            if skip_empty and not any(c.strip() for c in cells):
                continue
            rows.append(cells)
        return rows

    def _wide_as_blocks(self, rows: list[list[str]]) -> str:
        """Compact representation for wide sheets (better for LLM context)."""
        if not rows:
            return ""
        headers = rows[0]
        # If first row is empty-ish, synthesize headers.
        if not any(h.strip() for h in headers):
            headers = [f"Col{i + 1}" for i in range(len(headers))]
            data = rows[1:] if len(rows) > 1 else []
        else:
            data = rows[1:]

        blocks: list[str] = []
        for idx, row in enumerate(data, start=1):
            lines = [f"### Row {idx}"]
            for i, header in enumerate(headers):
                value = row[i] if i < len(row) else ""
                if not value.strip():
                    continue
                key = header.strip() or f"Col{i + 1}"
                lines.append(f"- **{key}**: {value}")
            if len(lines) > 1:
                blocks.append("\n".join(lines))
        if not blocks:
            # Only header or empty data — fall back to table.
            return gfm_table(rows, header=True) + "\n"
        return "\n\n".join(blocks) + "\n"


def _coerce_value(value: Any) -> Any:
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return value
    return value


def _trim_trailing_columns(rows: list[list[str]]) -> list[list[str]]:
    max_useful = 0
    for row in rows:
        for i in range(len(row) - 1, -1, -1):
            if row[i].strip():
                max_useful = max(max_useful, i + 1)
                break
    if max_useful == 0:
        return []
    return [row[:max_useful] for row in rows]
